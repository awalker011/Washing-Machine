import csv
import json

import pytest
from openpyxl import Workbook, load_workbook

from data_standardizer.pipeline import process_all


def test_duplicate_log_includes_rows_that_are_invalid_for_other_reasons(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"
    output_dir = tmp_path / "output"
    logs_dir = tmp_path / "logs"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (input_dir / "accounts.csv").write_text(
        "Account Name*,Account Type\n"
        "Acme Corp,???\n"
        "Acme Corp,???\n",
        encoding="utf-8",
    )

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*", "Account Type"],
                "fields": {
                    "Account Name*": {"type": "string", "required": True},
                    "Account Type": {"type": "string", "disallow_values": ["???"]},
                },
                "duplicate_rules": [
                    {
                        "name": "account_name_exact",
                        "type": "exact",
                        "fields": ["Account Name*"],
                        "exclude": True,
                        "message": "Duplicate Account Name* values were found.",
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    process_all(
        input_path=str(input_dir),
        schema_path=str(schema_dir),
        mapping_path=str(mapping_dir),
        output_dir=str(output_dir),
        logs_dir=str(logs_dir),
    )

    duplicate_log_path = logs_dir / "duplicate_log.csv"
    with duplicate_log_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    assert len(rows) == 2
    assert {row["row_number"] for row in rows} == {"2", "3"}


def test_process_all_writes_corrections_workbook_with_only_rejected_rows(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"
    output_dir = tmp_path / "output"
    logs_dir = tmp_path / "logs"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (input_dir / "accounts.csv").write_text(
        "Account Name*,Account Type\n"
        "Acme Corp,Retail\n"
        ",???\n",
        encoding="utf-8",
    )

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*", "Account Type"],
                "fields": {
                    "Account Name*": {"type": "string", "required": True},
                    "Account Type": {"type": "string", "disallow_values": ["???"]},
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = process_all(
        input_path=str(input_dir),
        schema_path=str(schema_dir),
        mapping_path=str(mapping_dir),
        output_dir=str(output_dir),
        logs_dir=str(logs_dir),
    )

    corrections_path = output_dir / "Needs_Correction.xlsx"
    assert result["corrections_file"] == str(corrections_path)
    assert corrections_path.exists()

    workbook = load_workbook(corrections_path)
    assert workbook.sheetnames == ["Accounts"]

    worksheet = workbook["Accounts"]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("Account Name*", "Account Type", "Source File", "Row Number", "Issues Found")
    assert len(rows) == 2

    data_row = rows[1]
    assert data_row[0] in (None, "")
    assert data_row[1] == "???"
    assert data_row[2] == "accounts.csv"
    assert data_row[3] == 3
    assert "Field is required" in data_row[4]


def test_process_all_rejects_rows_with_malformed_account_email(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"
    output_dir = tmp_path / "output"
    logs_dir = tmp_path / "logs"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (input_dir / "accounts.csv").write_text(
        "Account Name*,Account Email\n"
        "Acme Corp,jane@example.com\n"
        "Other Corp,not-an-email\n",
        encoding="utf-8",
    )

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*", "Account Email"],
                "fields": {
                    "Account Name*": {"type": "string", "required": True},
                    "Account Email": {"type": "string", "required": False, "format": "email"},
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = process_all(
        input_path=str(input_dir),
        schema_path=str(schema_dir),
        mapping_path=str(mapping_dir),
        output_dir=str(output_dir),
        logs_dir=str(logs_dir),
    )

    assert result["totals"] == {"rows_read": 2, "rows_accepted": 1, "rows_rejected": 1}
    assert result["corrections_file"] is not None

    with (output_dir / "Accounts.csv").open("r", encoding="utf-8", newline="") as handle:
        output_rows = list(csv.DictReader(handle))
    assert [row["Account Name*"] for row in output_rows] == ["Acme Corp"]

    with (output_dir / "Accounts_errors.csv").open("r", encoding="utf-8", newline="") as handle:
        error_rows = list(csv.DictReader(handle))
    assert len(error_rows) == 1
    assert error_rows[0]["field"] == "Account Email"
    assert error_rows[0]["severity"] == "blocker"


def test_process_all_omits_corrections_workbook_when_nothing_is_rejected(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"
    output_dir = tmp_path / "output"
    logs_dir = tmp_path / "logs"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (input_dir / "accounts.csv").write_text(
        "Account Name*\nAcme Corp\n",
        encoding="utf-8",
    )

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*"],
                "fields": {"Account Name*": {"type": "string", "required": True}},
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = process_all(
        input_path=str(input_dir),
        schema_path=str(schema_dir),
        mapping_path=str(mapping_dir),
        output_dir=str(output_dir),
        logs_dir=str(logs_dir),
    )

    assert result["corrections_file"] is None
    assert not (output_dir / "Needs_Correction.xlsx").exists()


def test_process_all_raises_when_no_input_files_match(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*"],
                "fields": {"Account Name*": {"type": "string", "required": True}},
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["missing.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    with pytest.raises(FileNotFoundError, match="No input files matched"):
        process_all(
            input_path=str(input_dir),
            schema_path=str(schema_dir),
            mapping_path=str(mapping_dir),
            output_dir=str(tmp_path / "output"),
            logs_dir=str(tmp_path / "logs"),
        )


def test_process_all_raises_when_configured_sheet_is_missing(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Wrong Sheet"
    worksheet.append(["Account Name*"])
    worksheet.append(["Acme Corp"])
    workbook.save(input_dir / "accounts.xlsx")

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*"],
                "fields": {"Account Name*": {"type": "string", "required": True}},
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.xlsx"],
                "sheet_name": "Accounts",
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="Worksheet 'Accounts' was not found"):
        process_all(
            input_path=str(input_dir),
            schema_path=str(schema_dir),
            mapping_path=str(mapping_dir),
            output_dir=str(tmp_path / "output"),
            logs_dir=str(tmp_path / "logs"),
        )


def test_process_all_normalizes_state_header_variants(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"
    output_dir = tmp_path / "output"
    logs_dir = tmp_path / "logs"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (input_dir / "accounts.csv").write_text(
        "Address 1: Street 1,Address 1: City,Address 1: State,Address 1: ZIP/Postal Code\n"
        "100 Main St,Chadds Ford,PA,19317\n",
        encoding="utf-8",
    )

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": [
                    "Address 1: Street 1",
                    "Address 1: City",
                    "Address 1: State/Province",
                    "Address 1: ZIP/Postal Code",
                ],
                "fields": {
                    "Address 1: Street 1": {"type": "string", "required": False},
                    "Address 1: City": {"type": "string", "required": False},
                    "Address 1: State/Province": {"type": "string", "required": False},
                    "Address 1: ZIP/Postal Code": {"type": "string", "required": False},
                },
                "cross_field_rules": [
                    {
                        "type": "required_if_any",
                        "trigger_fields": [
                            "Address 1: City",
                            "Address 1: State/Province",
                            "Address 1: ZIP/Postal Code",
                        ],
                        "required_fields": [
                            "Address 1: Street 1",
                            "Address 1: City",
                            "Address 1: State/Province",
                        ],
                        "message": "If structured account address data is provided, Street 1, City, and State/Province must all be populated.",
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    process_all(
        input_path=str(input_dir),
        schema_path=str(schema_dir),
        mapping_path=str(mapping_dir),
        output_dir=str(output_dir),
        logs_dir=str(logs_dir),
    )

    with (output_dir / "Accounts_errors.csv").open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    assert rows == []


def test_process_all_enriches_building_customer_without_legacy_id(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"
    output_dir = tmp_path / "output"
    logs_dir = tmp_path / "logs"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (input_dir / "accounts.csv").write_text(
        "Account Name*,Legacy Customer #\n"
        "Acme Co:123 Main St,QB-1\n",
        encoding="utf-8",
    )

    (input_dir / "building_locations.csv").write_text(
        "Name,Customer\n"
        "123 Main St,Acme.Co\n",
        encoding="utf-8",
    )

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*", "Legacy Customer #"],
                "fields": {
                    "Account Name*": {"type": "string", "required": True},
                    "Legacy Customer #": {"type": "string", "required": False},
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (schema_dir / "building_locations.json").write_text(
        json.dumps(
            {
                "entity_name": "Building Locations",
                "output_columns": ["Name*", "Customer (Account - for Invoicing)"],
                "fields": {
                    "Name*": {"type": "string", "required": True},
                    "Customer (Account - for Invoicing)": {"type": "string", "required": False},
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "building_locations.json").write_text(
        json.dumps(
            {
                "entity_name": "Building Locations",
                "schema": "Building Locations",
                "source_patterns": ["building_locations.csv"],
                "target_file": "Building Locations.csv",
                "error_file": "Building Locations_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    process_all(
        input_path=str(input_dir),
        schema_path=str(schema_dir),
        mapping_path=str(mapping_dir),
        output_dir=str(output_dir),
        logs_dir=str(logs_dir),
    )

    with (output_dir / "Building Locations.csv").open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    assert len(rows) == 1
    assert rows[0]["Customer (Account - for Invoicing)"] == "Acme Co:123 Main St"


def test_process_all_logs_building_customer_source_gap(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"
    output_dir = tmp_path / "output"
    logs_dir = tmp_path / "logs"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (input_dir / "accounts.csv").write_text(
        "Account Name*\n"
        "Acme Co\n",
        encoding="utf-8",
    )

    (input_dir / "building_locations.csv").write_text(
        "Name\n"
        "123 Main St\n",
        encoding="utf-8",
    )

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*"],
                "fields": {"Account Name*": {"type": "string", "required": True}},
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (schema_dir / "building_locations.json").write_text(
        json.dumps(
            {
                "entity_name": "Building Locations",
                "output_columns": ["Name*", "Customer (Account - for Invoicing)"],
                "fields": {
                    "Name*": {"type": "string", "required": True},
                    "Customer (Account - for Invoicing)": {"type": "string", "required": False},
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "building_locations.json").write_text(
        json.dumps(
            {
                "entity_name": "Building Locations",
                "schema": "Building Locations",
                "source_patterns": ["building_locations.csv"],
                "target_file": "Building Locations.csv",
                "error_file": "Building Locations_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    process_all(
        input_path=str(input_dir),
        schema_path=str(schema_dir),
        mapping_path=str(mapping_dir),
        output_dir=str(output_dir),
        logs_dir=str(logs_dir),
    )

    with (output_dir / "Building Locations_errors.csv").open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    assert len(rows) == 1
    assert rows[0]["field"] == "Customer (Account - for Invoicing)"
    assert "source gap" in rows[0]["reason"]


def test_process_all_logs_building_customer_non_unique_match(tmp_path):
    input_dir = tmp_path / "input"
    schema_dir = tmp_path / "schemas"
    mapping_dir = tmp_path / "mappings"
    output_dir = tmp_path / "output"
    logs_dir = tmp_path / "logs"

    input_dir.mkdir()
    schema_dir.mkdir()
    mapping_dir.mkdir()

    (input_dir / "accounts.csv").write_text(
        "Account Name*\n"
        "Acme Co\n"
        "ACME-CO\n",
        encoding="utf-8",
    )

    (input_dir / "building_locations.csv").write_text(
        "Name,Customer\n"
        "123 Main St,Acme.Co\n",
        encoding="utf-8",
    )

    (schema_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "output_columns": ["Account Name*"],
                "fields": {"Account Name*": {"type": "string", "required": True}},
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (schema_dir / "building_locations.json").write_text(
        json.dumps(
            {
                "entity_name": "Building Locations",
                "output_columns": ["Name*", "Customer (Account - for Invoicing)"],
                "fields": {
                    "Name*": {"type": "string", "required": True},
                    "Customer (Account - for Invoicing)": {"type": "string", "required": False},
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "accounts.json").write_text(
        json.dumps(
            {
                "entity_name": "Accounts",
                "schema": "Accounts",
                "source_patterns": ["accounts.csv"],
                "target_file": "Accounts.csv",
                "error_file": "Accounts_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    (mapping_dir / "building_locations.json").write_text(
        json.dumps(
            {
                "entity_name": "Building Locations",
                "schema": "Building Locations",
                "source_patterns": ["building_locations.csv"],
                "target_file": "Building Locations.csv",
                "error_file": "Building Locations_errors.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    process_all(
        input_path=str(input_dir),
        schema_path=str(schema_dir),
        mapping_path=str(mapping_dir),
        output_dir=str(output_dir),
        logs_dir=str(logs_dir),
    )

    with (output_dir / "Building Locations_errors.csv").open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    assert len(rows) == 1
    assert rows[0]["field"] == "Customer (Account - for Invoicing)"
    assert "non-unique Customer name" in rows[0]["reason"]
