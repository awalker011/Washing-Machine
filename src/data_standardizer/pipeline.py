from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    load_workbook = None
    Font = None

from .config_loader import load_named_configs
from .entity_rules import evaluate_entity_rules
from .transformers import apply_transformations, is_blank, normalize_loose_text
from .validators import comparable, validate_record, validate_relationship_rules

ERROR_LOG_FIELDS = [
    "timestamp",
    "entity",
    "source_file",
    "row_number",
    "field",
    "severity",
    "dependency_type",
    "depends_on_entity",
    "depends_on_source_file",
    "depends_on_row_number",
    "suggested_fix",
    "reason",
    "original_value",
    "raw_row",
]
RUN_LOG_FIELDS = [
    "timestamp",
    "entity",
    "files_processed",
    "rows_read",
    "rows_accepted",
    "rows_rejected",
    "duplicate_rows_flagged",
    "output_file",
    "error_file",
]
DUPLICATE_LOG_FIELDS = [
    "timestamp",
    "entity",
    "duplicate_type",
    "rule_name",
    "key_fields",
    "key_value",
    "source_file",
    "row_number",
    "action",
    "reason",
    "related_rows",
    "raw_row",
]

ProgressCallback = Callable[[dict[str, Any]], None]


def process_all(
    input_path: str,
    schema_path: str,
    mapping_path: str,
    output_dir: str,
    logs_dir: str,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, Any]:
    input_root = Path(input_path)
    output_root = Path(output_dir)
    logs_root = Path(logs_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    logs_root.mkdir(parents=True, exist_ok=True)

    schemas = load_named_configs(schema_path)
    mappings = load_named_configs(mapping_path)
    if not mappings:
        raise ValueError("No mapping configuration files were found.")

    staged_entities: dict[str, dict[str, Any]] = {}
    relationship_index: dict[tuple[str, str], set[str]] = defaultdict(set)
    entity_row_index: dict[tuple[str, str], dict[str, list[dict[str, Any]]]] = defaultdict(
        lambda: defaultdict(list)
    )
    total_rows_read = 0
    total_rows_accepted = 0
    total_rows_rejected = 0
    entity_count = len(mappings)

    _emit_progress(
        progress_callback,
        phase="starting",
        message="Loading mappings and preparing run.",
        current=0,
        total=max(1, entity_count * 2 + 3),
    )

    for stage_index, (entity_name, mapping) in enumerate(mappings.items(), start=1):
        _emit_progress(
            progress_callback,
            phase="staging",
            message=f"Staging {entity_name}...",
            entity=entity_name,
            current=stage_index,
            total=max(1, entity_count * 2 + 3),
        )
        schema = _resolve_schema(entity_name, mapping, schemas)
        staged_rows, all_rows, entity_errors, processed_files, read_count = _stage_entity(
            input_root=input_root,
            entity_name=entity_name,
            mapping=mapping,
            schema=schema,
        )
        total_rows_read += read_count

        for staged_row in staged_rows:
            for field_name, value in staged_row["data"].items():
                normalized = comparable(value)
                if normalized is not None:
                    relationship_index[(entity_name, field_name)].add(normalized)
                    entity_row_index[(entity_name, field_name)][normalized].append(staged_row)

        staged_entities[entity_name] = {
            "schema": schema,
            "mapping": mapping,
            "rows": staged_rows,
            "all_rows": all_rows,
            "errors": entity_errors,
            "processed_files": processed_files,
            "rows_read": read_count,
        }

    rejected_reference_index = _build_rejected_reference_index(staged_entities)

    _emit_progress(
        progress_callback,
        phase="enrichment",
        message="Applying cross-sheet enrichment.",
        current=entity_count + 1,
        total=max(1, entity_count * 2 + 3),
    )
    _enrich_cross_sheet_links(staged_entities)

    entity_rule_issues = evaluate_entity_rules(staged_entities)
    entity_summaries: list[dict[str, Any]] = []
    duplicate_log_rows: list[dict[str, Any]] = []
    corrections_by_entity: dict[str, list[dict[str, Any]]] = {}

    for validation_index, (entity_name, entity_state) in enumerate(staged_entities.items(), start=1):
        _emit_progress(
            progress_callback,
            phase="validating",
            message=f"Validating {entity_name}...",
            entity=entity_name,
            current=entity_count + 1 + validation_index,
            total=max(1, entity_count * 2 + 3),
        )
        schema = entity_state["schema"]
        mapping = entity_state["mapping"]
        final_rows: list[dict[str, Any]] = []
        errors = list(entity_state["errors"])

        duplicate_exclusions, entity_duplicate_logs = _detect_duplicates(
            entity_name=entity_name,
            staged_rows=entity_state.get("all_rows", entity_state["rows"]),
            schema=schema,
        )
        duplicate_log_rows.extend(entity_duplicate_logs)

        for staged_row in entity_state["rows"]:
            row_key = (staged_row["source_file"], staged_row["row_number"])
            if row_key in duplicate_exclusions:
                continue

            row_issues = list(
                entity_rule_issues.get((entity_name, staged_row["source_file"], staged_row["row_number"]), [])
            )
            row_issues.extend(staged_row.get("enrichment_issues", []))
            row_issues.extend(
                validate_relationship_rules(
                    staged_row["data"],
                    schema.get("relationship_rules", []),
                    relationship_index,
                    entity_row_index,
                    rejected_reference_index,
                )
            )
            if row_issues:
                has_blocker = _append_row_issues(
                    errors=errors,
                    entity_name=entity_name,
                    source_file=staged_row["source_file"],
                    row_number=staged_row["row_number"],
                    issues=row_issues,
                    record=staged_row["data"],
                    raw_row=staged_row["raw_row"],
                )
                if not has_blocker:
                    final_rows.append(staged_row["data"])
            else:
                final_rows.append(staged_row["data"])

        output_columns = schema.get("output_columns") or list(schema.get("fields", {}).keys())
        target_file = mapping.get("target_file", f"{entity_name}_standardized.csv")
        error_file = mapping.get("error_file", f"{entity_name}_errors.csv")

        _write_csv(output_root / target_file, final_rows, output_columns)
        _write_csv(output_root / error_file, errors, ERROR_LOG_FIELDS)

        _extend_rejected_reference_index_for_entity(
            rejected_reference_index,
            entity_name=entity_name,
            staged_rows=entity_state.get("all_rows", entity_state["rows"]),
            errors=errors,
        )

        accepted = len(final_rows)
        rejected_keys = {
            (error["source_file"], error["row_number"])
            for error in errors
            if error.get("severity", "blocker") == "blocker"
        }
        rejected_row_keys = rejected_keys | duplicate_exclusions
        rejected = len(rejected_row_keys)
        duplicate_rows_flagged = len(
            {(row["source_file"], row["row_number"]) for row in entity_duplicate_logs}
        )
        total_rows_accepted += accepted
        total_rows_rejected += rejected

        if rejected_row_keys:
            entity_corrections = _collect_corrections_rows(
                staged_rows=entity_state.get("all_rows", entity_state["rows"]),
                errors=errors,
                entity_duplicate_logs=entity_duplicate_logs,
                rejected_row_keys=rejected_row_keys,
            )
            if entity_corrections:
                corrections_by_entity[entity_name] = entity_corrections

        entity_summaries.append(
            {
                "entity": entity_name,
                "files_processed": entity_state["processed_files"],
                "rows_read": entity_state["rows_read"],
                "rows_accepted": accepted,
                "rows_rejected": rejected,
                "duplicate_rows_flagged": duplicate_rows_flagged,
                "output_file": str(output_root / target_file),
                "error_file": str(output_root / error_file),
            }
        )

    _append_run_log(logs_root / "run_log.csv", entity_summaries)
    _append_duplicate_log(logs_root / "duplicate_log.csv", duplicate_log_rows)
    customer_summary = _build_customer_summary(entity_summaries)
    customer_summary_path = output_root / "customer_summary.txt"
    customer_summary_path.write_text(customer_summary, encoding="utf-8")

    corrections_file_path: str | None = None
    if corrections_by_entity:
        corrections_workbook_path = output_root / "Needs_Correction.xlsx"
        _write_corrections_workbook(corrections_workbook_path, corrections_by_entity)
        corrections_file_path = str(corrections_workbook_path)

    _emit_progress(
        progress_callback,
        phase="complete",
        message="Run complete.",
        current=max(1, entity_count * 2 + 3),
        total=max(1, entity_count * 2 + 3),
    )

    return {
        "status": "completed",
        "entities": entity_summaries,
        "duplicate_log_file": str(logs_root / "duplicate_log.csv"),
        "customer_summary_file": str(customer_summary_path),
        "corrections_file": corrections_file_path,
        "totals": {
            "rows_read": total_rows_read,
            "rows_accepted": total_rows_accepted,
            "rows_rejected": total_rows_rejected,
        },
    }


def _enrich_cross_sheet_links(staged_entities: dict[str, dict[str, Any]]) -> None:
    """Apply best-effort cross-sheet link enrichment to reduce dropped relationship data."""
    _enrich_building_customer_from_accounts(staged_entities)


def _enrich_building_customer_from_accounts(staged_entities: dict[str, dict[str, Any]]) -> None:
    accounts_state = staged_entities.get("Accounts")
    buildings_state = staged_entities.get("Building Locations")
    if not accounts_state or not buildings_state:
        return

    accounts_rows = accounts_state.get("rows", [])
    building_rows = buildings_state.get("rows", [])
    if not accounts_rows or not building_rows:
        return

    legacy_to_account: dict[str, str] = {}
    exact_name_to_account: dict[str, str] = {}
    normalized_name_to_accounts: dict[str, set[str]] = defaultdict(set)

    for row in accounts_rows:
        data = row.get("data", {})
        account_name = comparable(data.get("Account Name*"))
        if account_name is None:
            continue

        legacy_id = comparable(data.get("Legacy Customer #"))
        if legacy_id is not None:
            legacy_to_account[legacy_id.casefold()] = account_name

        exact_name_to_account[account_name.casefold()] = account_name
        normalized_name = normalize_loose_text(account_name)
        if normalized_name:
            normalized_name_to_accounts[normalized_name].add(account_name)

    for row in building_rows:
        data = row.get("data", {})
        raw_row = row.get("raw_row", {})

        current_customer = comparable(data.get("Customer (Account - for Invoicing)"))
        if current_customer is not None and current_customer.casefold() in exact_name_to_account:
            # Canonicalize existing match to exact account casing.
            data["Customer (Account - for Invoicing)"] = exact_name_to_account[current_customer.casefold()]
            continue

        legacy_id = _get_raw_value(raw_row, [
            "Legacy Account Id",
            "Legacy Customer #",
            "Customer Legacy ID",
            "QB Legacy ID",
        ])
        raw_customer = _get_raw_value(raw_row, [
            "Customer",
            "Company Name",
            "Account Name",
        ])
        building_name = _get_raw_value(raw_row, ["Name", "Building Name"])

        resolved_account: str | None = None
        ambiguous_customer_matches: set[str] = set()
        ambiguous_composite_matches: set[str] = set()

        if legacy_id:
            resolved_account = legacy_to_account.get(legacy_id.casefold())

        if resolved_account is None and raw_customer:
            resolved_account = exact_name_to_account.get(raw_customer.casefold())

        if resolved_account is None and raw_customer:
            normalized_customer = normalize_loose_text(raw_customer)
            customer_matches = normalized_name_to_accounts.get(normalized_customer, set())
            if len(customer_matches) == 1:
                resolved_account = next(iter(customer_matches))
            elif len(customer_matches) > 1:
                ambiguous_customer_matches = set(customer_matches)

        # Fallback: some sources store account as "<customer>:<building name>".
        if resolved_account is None and raw_customer and building_name:
            composite = f"{raw_customer}:{building_name}"
            composite_match = exact_name_to_account.get(composite.casefold())
            if composite_match is not None:
                resolved_account = composite_match
            else:
                normalized_composite = normalize_loose_text(composite)
                composite_matches = normalized_name_to_accounts.get(normalized_composite, set())
                if len(composite_matches) == 1:
                    resolved_account = next(iter(composite_matches))
                elif len(composite_matches) > 1:
                    ambiguous_composite_matches = set(composite_matches)

        if resolved_account is not None:
            data["Customer (Account - for Invoicing)"] = resolved_account
            continue

        field_name = "Customer (Account - for Invoicing)"
        if legacy_id is None and raw_customer is None:
            _add_enrichment_issue(
                row,
                field_name,
                "Unable to pair building to an account: source gap (missing Customer and Legacy Account Id).",
            )
            continue

        if ambiguous_customer_matches:
            candidates = ", ".join(sorted(ambiguous_customer_matches)[:5])
            _add_enrichment_issue(
                row,
                field_name,
                (
                    "Unable to pair building to an account: non-unique Customer name matched multiple Accounts "
                    f"({candidates})."
                ),
            )
            continue

        if ambiguous_composite_matches:
            candidates = ", ".join(sorted(ambiguous_composite_matches)[:5])
            _add_enrichment_issue(
                row,
                field_name,
                (
                    "Unable to pair building to an account: non-unique Customer+Building composite matched multiple Accounts "
                    f"({candidates})."
                ),
            )
            continue

        detail_parts = []
        if legacy_id is not None:
            detail_parts.append(f"Legacy Account Id='{legacy_id}'")
        if raw_customer is not None:
            detail_parts.append(f"Customer='{raw_customer}'")
        if building_name is not None:
            detail_parts.append(f"Building Name='{building_name}'")
        details = "; ".join(detail_parts) if detail_parts else "no source values"
        _add_enrichment_issue(
            row,
            field_name,
            f"Unable to pair building to an account: no matching Account found ({details}).",
        )


def _add_enrichment_issue(row: dict[str, Any], field_name: str, reason: str) -> None:
    issues = row.setdefault("enrichment_issues", [])
    issues.append((field_name, reason))


def _get_raw_value(raw_row: dict[str, Any], candidates: list[str]) -> str | None:
    if not raw_row:
        return None

    normalized_lookup = {
        _normalize_header(key, case_sensitive=False): value
        for key, value in raw_row.items()
        if key is not None
    }

    for candidate in candidates:
        candidate_key = _normalize_header(candidate, case_sensitive=False)
        value = normalized_lookup.get(candidate_key)
        normalized = comparable(value)
        if normalized is not None:
            return normalized
    return None


def _resolve_schema(entity_name: str, mapping: dict, schemas: dict[str, dict]) -> dict:
    schema_name = mapping.get("schema") or entity_name
    schema = schemas.get(schema_name)
    if schema is None:
        raise ValueError(f"No schema found for entity '{entity_name}' (expected key '{schema_name}').")
    return schema


def _stage_entity(
    input_root: Path,
    entity_name: str,
    mapping: dict,
    schema: dict,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[str], int]:
    staged_rows: list[dict[str, Any]] = []
    all_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    processed_files: list[str] = []
    rows_read = 0

    input_files = _find_input_files(input_root, mapping.get("source_patterns"))
    if not input_files:
        patterns = list(mapping.get("source_patterns") or ["*.csv", "*.xlsx"])
        raise FileNotFoundError(
            f"No input files matched for entity '{entity_name}' under '{input_root}' using patterns: {patterns}"
        )

    for source_file in input_files:
        processed_files.append(source_file.name)
        for row_number, raw_row in _iter_input_rows(source_file, mapping.get("sheet_name")):
            rows_read += 1
            standardized = _map_row(raw_row, mapping, schema)
            transformed, transform_issues = apply_transformations(
                standardized,
                mapping.get("transformations", {}),
            )

            row_state = {
                "data": transformed,
                "source_file": source_file.name,
                "row_number": row_number,
                "raw_row": raw_row,
            }
            all_rows.append(row_state)

            if transform_issues:
                has_blocker = _append_row_issues(
                    errors=errors,
                    entity_name=entity_name,
                    source_file=source_file.name,
                    row_number=row_number,
                    issues=transform_issues,
                    record=standardized,
                    raw_row=raw_row,
                )
                if has_blocker:
                    continue

            row_issues = validate_record(transformed, schema)
            if row_issues:
                has_blocker = _append_row_issues(
                    errors=errors,
                    entity_name=entity_name,
                    source_file=source_file.name,
                    row_number=row_number,
                    issues=row_issues,
                    record=transformed,
                    raw_row=raw_row,
                )
                if has_blocker:
                    continue

            staged_rows.append(row_state)

    return staged_rows, all_rows, errors, processed_files, rows_read


def _find_input_files(input_root: Path, source_patterns: Iterable[str] | None) -> list[Path]:
    if input_root.is_file():
        return [input_root]

    patterns = list(source_patterns or ["*.csv", "*.xlsx"])
    matches: list[Path] = []

    for pattern in patterns:
        matches.extend(path for path in input_root.glob(pattern) if path.is_file())

    deduplicated = sorted({path.resolve() for path in matches})
    return [Path(path) for path in deduplicated]


def _iter_input_rows(file_path: Path, sheet_name: str | int | None):
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        yield from _iter_csv_rows(file_path)
        return

    if suffix == ".xlsx":
        yield from _iter_xlsx_rows(file_path, sheet_name)
        return

    raise ValueError(f"Unsupported input file type: {file_path}")


def _clean_header_label(value: Any) -> str:
    """Tidy whitespace/newlines in a header for display and dict-key use, preserving original casing."""
    if value is None:
        return ""

    text = str(value).strip()
    text = re.sub(r"[\r\n\t]+", " ", text)
    return re.sub(r" +", " ", text).strip()


def _iter_csv_rows(file_path: Path):
    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"CSV file '{file_path.name}' is empty or missing a header row.")

        reader.fieldnames = [_clean_header_label(name) for name in reader.fieldnames]
        if not any(reader.fieldnames):
            raise ValueError(f"CSV file '{file_path.name}' does not contain any usable column headers.")

        for row_number, row in enumerate(reader, start=2):
            cleaned = {
                _clean_header_label(key): value
                for key, value in row.items()
                if key is not None and _clean_header_label(key)
            }
            if any(not is_blank(value) for value in cleaned.values()):
                yield row_number, cleaned


def _iter_xlsx_rows(file_path: Path, sheet_name: str | int | None):
    if load_workbook is None:
        raise RuntimeError("XLSX support requires openpyxl to be installed.")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            worksheet = workbook.worksheets[0]
        elif isinstance(sheet_name, int):
            if sheet_name < 0 or sheet_name >= len(workbook.worksheets):
                available_sheets = ", ".join(workbook.sheetnames)
                raise ValueError(
                    f"Worksheet index {sheet_name} is out of range for '{file_path.name}'. Available sheets: {available_sheets}"
                )
            worksheet = workbook.worksheets[sheet_name]
        else:
            if str(sheet_name) not in workbook.sheetnames:
                available_sheets = ", ".join(workbook.sheetnames)
                raise ValueError(
                    f"Worksheet '{sheet_name}' was not found in '{file_path.name}'. Available sheets: {available_sheets}"
                )
            worksheet = workbook[str(sheet_name)]

        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            raise ValueError(
                f"Worksheet '{worksheet.title}' in '{file_path.name}' is empty or missing a header row."
            )

        headers = [_clean_header_label(value) for value in header_row]
        if not any(headers):
            raise ValueError(
                f"Worksheet '{worksheet.title}' in '{file_path.name}' does not contain any usable column headers."
            )

        for row_number, values in enumerate(rows, start=2):
            cleaned = {}
            for header, value in zip(headers, values):
                if header:
                    cleaned[header] = value
            if any(not is_blank(value) for value in cleaned.values()):
                yield row_number, cleaned
    finally:
        workbook.close()


def _map_row(raw_row: dict[str, Any], mapping: dict, schema: dict) -> dict[str, Any]:
    output_columns = schema.get("output_columns") or list(schema.get("fields", {}).keys())
    field_mappings = mapping.get("field_mappings", {})
    mapping_defaults = mapping.get("defaults", {})
    schema_fields = schema.get("fields", {})

    standardized = {column: None for column in output_columns}
    
    # Build case-insensitive lookup for raw_row
    raw_row_lookup = _build_header_lookup(list(raw_row.keys()), case_sensitive=False)

    # Apply explicit field mappings.
    # Supports either:
    #   "Source Header": "Target Field"
    # or
    #   "Source Header": ["Target Field A", "Target Field B"]
    for source_column, target_field in field_mappings.items():
        normalized_source = _normalize_header(source_column, case_sensitive=False)
        matched_source = raw_row_lookup.get(normalized_source)
        if not matched_source:
            continue

        source_value = raw_row.get(matched_source)
        target_fields = target_field if isinstance(target_field, list) else [target_field]
        for mapped_target in target_fields:
            if mapped_target in standardized and is_blank(standardized.get(mapped_target)):
                standardized[mapped_target] = source_value

    # Apply column matching with fallback strategies
    for target_field in output_columns:
        if not is_blank(standardized.get(target_field)):
            continue  # Already found value via field mapping
            
        # Strategy 1: Direct key match with normalization
        normalized_target = _normalize_header(target_field, case_sensitive=False)
        matched = raw_row_lookup.get(normalized_target)
        if matched:
            standardized[target_field] = raw_row.get(matched)
            continue

        # Strategy 2: Try without trailing asterisk (for required field indicators)
        if target_field.endswith('*'):
            field_without_asterisk = target_field[:-1]
            normalized_without_asterisk = _normalize_header(field_without_asterisk, case_sensitive=False)
            matched = raw_row_lookup.get(normalized_without_asterisk)
            if matched:
                standardized[target_field] = raw_row.get(matched)
                continue

        # Strategy 3: Use mapping defaults
        if target_field in mapping_defaults:
            standardized[target_field] = mapping_defaults[target_field]
            continue

        # Strategy 4: Use schema defaults
        field_config = schema_fields.get(target_field, {})
        if "default" in field_config:
            standardized[target_field] = field_config["default"]

    return standardized


def _normalize_header(value: Any, *, case_sensitive: bool = False) -> str:
    """
    Normalize a header name for consistent matching.
    
    Handles:
    - None/empty values
    - Leading/trailing whitespace
    - Internal whitespace (multiple spaces, tabs, newlines)
    - Case normalization (optional)
    
    Args:
        value: The header name to normalize
        case_sensitive: If False, convert to lowercase for matching
    
    Returns:
        Normalized header string
    """
    if value is None:
        return ""
    
    # Convert to string and handle common separators
    text = str(value).strip()
    
    # Replace various whitespace characters (newlines, tabs, etc.) with single space
    text = re.sub(r'[\r\n\t]+', ' ', text)
    
    # Collapse multiple spaces into single space
    text = re.sub(r' +', ' ', text).strip()

    # Canonicalize common State/Province variants so mappings stay consistent.
    text = re.sub(r"\s*/\s*", "/", text)
    text = re.sub(r"\bprovince/state\b", "state/province", text, flags=re.IGNORECASE)
    text = re.sub(
        r"\b(address\s*\d*\s*:\s*)state\b(?!\s*/\s*province)",
        r"\1state/province",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r":\s*state$", ": state/province", text, flags=re.IGNORECASE)
    if re.fullmatch(r"state", text, flags=re.IGNORECASE):
        text = "state/province"
    
    # Optionally normalize case
    if not case_sensitive:
        text = text.casefold()
    
    return text


def _build_header_lookup(headers: list[str], case_sensitive: bool = False) -> dict[str, str]:
    """
    Build a lookup map from normalized headers to original headers.
    
    Helps identify cases where headers differ only by normalization issues.
    """
    lookup = {}
    for header in headers:
        normalized = _normalize_header(header, case_sensitive=case_sensitive)
        if normalized:
            lookup[normalized] = header
    return lookup


def _find_matching_headers(source_headers: list[str], expected_headers: list[str], case_sensitive: bool = False) -> dict[str, str | None]:
    """
    Match source headers to expected headers, handling normalization.
    
    Returns:
        Dict mapping expected header names to found source headers (or None if not found)
    """
    source_lookup = _build_header_lookup(source_headers, case_sensitive=case_sensitive)
    matches = {}
    
    for expected in expected_headers:
        normalized_expected = _normalize_header(expected, case_sensitive=case_sensitive)
        matches[expected] = source_lookup.get(normalized_expected)
    
    return matches


def _validate_columns(
    source_headers: list[str],
    expected_headers: list[str] | None,
    entity_name: str,
    file_name: str,
    case_sensitive: bool = False,
) -> tuple[dict[str, str | None], list[str]]:
    """
    Validate that expected columns exist in source headers.
    
    Returns:
        Tuple of (matches dict, list of warnings)
    """
    warnings = []
    
    if not expected_headers:
        return {}, warnings
    
    matches = _find_matching_headers(source_headers, expected_headers, case_sensitive=case_sensitive)
    
    # Find missing required columns
    missing = [h for h, match in matches.items() if match is None and h.endswith('*')]
    if missing:
        warnings.append(
            f"Entity '{entity_name}' in '{file_name}': "
            f"Missing required columns: {', '.join(missing)}"
        )
    
    # Find unmapped source columns (optional warning)
    matched_sources = set(m for m in matches.values() if m is not None)
    normalized_sources = set(_normalize_header(h, case_sensitive=case_sensitive) for h in source_headers)
    matched_normalized = set(_normalize_header(h, case_sensitive=case_sensitive) for h in matched_sources)
    unmapped = normalized_sources - matched_normalized
    
    if unmapped:
        warnings.append(
            f"Entity '{entity_name}' in '{file_name}': "
            f"Source columns not used in schema: {', '.join(unmapped)}"
        )
    
    return matches, warnings


def _detect_duplicates(
    entity_name: str,
    staged_rows: list[dict[str, Any]],
    schema: dict,
) -> tuple[set[tuple[str, int]], list[dict[str, Any]]]:
    duplicate_rules = schema.get("duplicate_rules", [])
    excluded_rows: set[tuple[str, int]] = set()
    log_rows: list[dict[str, Any]] = []

    for rule in duplicate_rules:
        fields = [str(field) for field in rule.get("fields", []) if str(field).strip()]
        if not fields:
            continue

        duplicate_type = str(rule.get("type", "exact"))
        rule_name = str(rule.get("name") or f"{duplicate_type}_{'_'.join(fields)}")
        normalization = str(rule.get("normalization") or ("strict" if duplicate_type == "exact" else "loose"))
        exclude = bool(rule.get("exclude", duplicate_type == "exact"))
        require_all_fields = bool(rule.get("require_all_fields", True))
        message = str(
            rule.get(
                "message",
                "Exact duplicate detected." if duplicate_type == "exact" else "Potential duplicate detected.",
            )
        )

        grouped_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for staged_row in staged_rows:
            key_parts = [
                _normalize_duplicate_value(staged_row["data"].get(field), normalization)
                for field in fields
            ]
            if require_all_fields and any(not part for part in key_parts):
                continue
            if not any(key_parts):
                continue
            grouped_rows[" | ".join(key_parts)].append(staged_row)

        for key_value, matches in grouped_rows.items():
            if len(matches) < 2:
                continue

            related_rows = "; ".join(f"{match['source_file']}:{match['row_number']}" for match in matches)
            action = "excluded_from_output" if exclude else "logged_only"

            for match in matches:
                row_key = (match["source_file"], match["row_number"])
                if exclude:
                    excluded_rows.add(row_key)

                log_rows.append(
                    {
                        "timestamp": datetime.now().isoformat(timespec="seconds"),
                        "entity": entity_name,
                        "duplicate_type": duplicate_type,
                        "rule_name": rule_name,
                        "key_fields": "; ".join(fields),
                        "key_value": key_value,
                        "source_file": match["source_file"],
                        "row_number": match["row_number"],
                        "action": action,
                        "reason": message,
                        "related_rows": related_rows,
                        "raw_row": _serialize_row(match["raw_row"]),
                    }
                )

    return excluded_rows, log_rows


def _normalize_duplicate_value(value: Any, normalization: str = "strict") -> str:
    if is_blank(value):
        return ""

    text = str(value).strip()
    if normalization == "strict":
        return text

    return normalize_loose_text(value, strip_punctuation=normalization == "loose")


def _append_row_issues(
    errors: list[dict[str, Any]],
    entity_name: str,
    source_file: str,
    row_number: int,
    issues: list[tuple[str, str]],
    record: dict[str, Any],
    raw_row: dict[str, Any],
) -> bool:
    has_blocker = False
    for field_name, reason in issues:
        error_row = _build_error_row(
            entity_name=entity_name,
            source_file=source_file,
            row_number=row_number,
            field_name=field_name,
            reason=reason,
            original_value=record.get(field_name),
            raw_row=raw_row,
        )
        if error_row.get("severity", "blocker") == "blocker":
            has_blocker = True
        errors.append(error_row)

    return has_blocker


def _emit_progress(progress_callback: ProgressCallback | None, **event: Any) -> None:
    if progress_callback is None:
        return
    progress_callback(event)


def _normalize_row_key(source_file: str, row_number: Any) -> tuple[str, int] | None:
    try:
        return source_file, int(row_number)
    except (TypeError, ValueError):
        return None


def _extend_rejected_reference_index_for_entity(
    rejected_reference_index: dict[tuple[str, str], dict[str, list[dict[str, Any]]]],
    entity_name: str,
    staged_rows: list[dict[str, Any]],
    errors: list[dict[str, Any]],
) -> None:
    reasons_by_row: dict[tuple[str, int], list[str]] = defaultdict(list)
    for error in errors:
        if error.get("severity", "blocker") != "blocker":
            continue
        source_file = str(error.get("source_file", "")).strip()
        row_key = _normalize_row_key(source_file, error.get("row_number"))
        if row_key is None:
            continue

        field_name = str(error.get("field", "")).strip()
        reason = str(error.get("reason", "")).strip()
        if field_name and reason:
            reasons_by_row[row_key].append(f"{field_name}: {reason}")

    if not reasons_by_row:
        return

    for row in staged_rows:
        row_key = (row["source_file"], row["row_number"])
        row_reasons = reasons_by_row.get(row_key)
        if not row_reasons:
            continue

        unique_reasons = list(dict.fromkeys(row_reasons))
        for field_name, value in row.get("data", {}).items():
            normalized = comparable(value)
            if normalized is None:
                continue

            rejected_reference_index.setdefault((entity_name, str(field_name)), {}).setdefault(normalized, []).append(
                {
                    "source_file": row["source_file"],
                    "row_number": row["row_number"],
                    "reasons": unique_reasons,
                }
            )


def _collect_corrections_rows(
    staged_rows: list[dict[str, Any]],
    errors: list[dict[str, Any]],
    entity_duplicate_logs: list[dict[str, Any]],
    rejected_row_keys: set[tuple[str, int]],
) -> list[dict[str, Any]]:
    reasons_by_row: dict[tuple[str, int], list[str]] = defaultdict(list)

    for error in errors:
        if error.get("severity", "blocker") != "blocker":
            continue
        row_key = _normalize_row_key(error.get("source_file", ""), error.get("row_number"))
        if row_key is None:
            continue
        field_name = str(error.get("field", "")).strip()
        reason = str(error.get("reason", "")).strip()
        if field_name and reason:
            reasons_by_row[row_key].append(f"{field_name}: {reason}")

    for duplicate_row in entity_duplicate_logs:
        if duplicate_row.get("action") != "excluded_from_output":
            continue
        row_key = _normalize_row_key(duplicate_row.get("source_file", ""), duplicate_row.get("row_number"))
        if row_key is None:
            continue
        rule_name = str(duplicate_row.get("rule_name", "")).strip()
        reason = str(duplicate_row.get("reason", "")).strip()
        label = f"Duplicate ({rule_name}): {reason}" if rule_name else f"Duplicate: {reason}"
        reasons_by_row[row_key].append(label)

    corrections_rows: list[dict[str, Any]] = []
    for row in staged_rows:
        row_key = (row["source_file"], row["row_number"])
        if row_key not in rejected_row_keys:
            continue

        corrections_rows.append(
            {
                "raw_row": row["raw_row"],
                "source_file": row["source_file"],
                "row_number": row["row_number"],
                "issues": list(dict.fromkeys(reasons_by_row.get(row_key, []))),
            }
        )

    return corrections_rows


def _safe_sheet_title(name: str) -> str:
    sanitized = re.sub(r"[:\\/?*\[\]]", " ", name).strip()
    sanitized = re.sub(r"\s+", " ", sanitized)
    return (sanitized or "Sheet")[:31]


def _write_corrections_workbook(file_path: Path, corrections_by_entity: dict[str, list[dict[str, Any]]]) -> None:
    if Workbook is None:
        raise RuntimeError("XLSX support requires openpyxl to be installed.")

    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for entity_name, rows in corrections_by_entity.items():
        headers: list[str] = []
        seen_headers: set[str] = set()
        for row in rows:
            for header in row["raw_row"].keys():
                if header not in seen_headers:
                    seen_headers.add(header)
                    headers.append(header)

        full_headers = [*headers, "Source File", "Row Number", "Issues Found"]
        worksheet = workbook.create_sheet(title=_safe_sheet_title(entity_name))
        worksheet.append(full_headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"

        for row in rows:
            raw_row = row["raw_row"]
            values = [raw_row.get(header) for header in headers]
            values.append(row["source_file"])
            values.append(row["row_number"])
            values.append("; ".join(row["issues"]) if row["issues"] else "")
            worksheet.append(values)

        for column_cells in worksheet.columns:
            longest = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(60, max(10, longest + 2))

    workbook.save(file_path)


def _build_customer_summary(entity_summaries: list[dict[str, Any]]) -> str:
    total_accepted = sum(int(entity.get("rows_accepted", 0)) for entity in entity_summaries)
    total_rejected = sum(int(entity.get("rows_rejected", 0)) for entity in entity_summaries)
    lines = [
        "Data Wash Summary",
        "",
        f"Accepted rows: {total_accepted}",
        f"Rejected rows: {total_rejected}",
        "",
        "Per entity:",
    ]
    for entity in entity_summaries:
        lines.append(
            (
                f"- {entity.get('entity')}: accepted {entity.get('rows_accepted', 0)}, "
                f"rejected {entity.get('rows_rejected', 0)}, duplicates flagged {entity.get('duplicate_rows_flagged', 0)}"
            )
        )
    lines.append("")
    lines.append("Use *_errors.csv files for row-level action details and suggested fixes.")
    return "\n".join(lines)


def _parse_dependency_metadata(reason: str) -> dict[str, str]:
    if "Lookup dependency note:" not in reason:
        return {
            "dependency_type": "primary",
            "depends_on_entity": "",
            "depends_on_source_file": "",
            "depends_on_row_number": "",
        }

    depends_on_entity = ""
    entity_match = re.search(r"referenced\s+(.+?)\s+row\s+was\s+rejected", reason)
    if entity_match:
        depends_on_entity = entity_match.group(1).strip()

    depends_on_source_file = ""
    depends_on_row_number = ""
    row_match = re.search(r"example\s+([^:()]+):(\d+)", reason)
    if row_match:
        depends_on_source_file = row_match.group(1).strip()
        depends_on_row_number = row_match.group(2).strip()

    return {
        "dependency_type": "downstream",
        "depends_on_entity": depends_on_entity,
        "depends_on_source_file": depends_on_source_file,
        "depends_on_row_number": depends_on_row_number,
    }


def _classify_error(
    *,
    field_name: str,
    reason: str,
    dependency_type: str,
) -> tuple[str, str, str]:
    reason_lower = reason.casefold()
    field_lower = field_name.casefold()

    if dependency_type == "downstream":
        return "REL-002", "warning", "Fix the upstream rejected record referenced in this message, then rerun."

    if "must exactly match" in reason_lower or "must exist" in reason_lower:
        if "similar existing value(s)" in reason_lower:
            return "REL-003", "blocker", "Use the suggested similar value and make names exactly match the reference sheet."
        return "REL-001", "blocker", "Add or correct the referenced parent value so this lookup resolves exactly."

    if "field is required" in reason_lower:
        return "VAL-001", "blocker", "Populate the required field with a non-empty value."

    if "expected value of type" in reason_lower:
        return "VAL-002", "blocker", "Use a value that matches the expected type for this field."

    if "plain contact name" in reason_lower:
        return "FMT-EMAIL", "warning", "Enter a valid email format (name@example.com), or leave as plain text if this is a contact name rather than an email."

    if field_lower == "account email" and "valid email" in reason_lower:
        return "FMT-EMAIL", "blocker", "Enter a valid email address (name@example.com) for Account Email before this row can be accepted."

    if "valid email" in reason_lower or "email" in field_lower:
        return "FMT-EMAIL", "warning", "Enter a valid email format (name@example.com) or leave blank if optional."

    if "valid phone" in reason_lower:
        return "FMT-PHONE", "warning", "Enter a phone number with at least 10 digits."

    if "postal" in reason_lower or "zip" in reason_lower:
        return "FMT-POSTAL", "warning", "Provide a valid US ZIP or Canadian postal code."

    if "disallowed placeholder" in reason_lower:
        return "VAL-003", "warning", "Replace placeholder text with a real business value."

    if "duplicate" in reason_lower:
        return "DUP-001", "blocker", "Resolve duplicate records so each unique key appears once."

    return "VAL-999", "blocker", "Review and correct the value based on the validation reason."


def _build_rejected_reference_index(
    staged_entities: dict[str, dict[str, Any]],
) -> dict[tuple[str, str], dict[str, list[dict[str, Any]]]]:
    index: dict[tuple[str, str], dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))

    for entity_name, entity_state in staged_entities.items():
        all_rows = entity_state.get("all_rows", [])
        entity_errors = entity_state.get("errors", [])

        blocker_row_keys: set[tuple[str, int]] = set()
        reasons_by_row: dict[tuple[str, int], list[str]] = defaultdict(list)
        for error in entity_errors:
            if error.get("severity", "blocker") != "blocker":
                continue
            source_file = str(error.get("source_file", "")).strip()
            row_number_raw = error.get("row_number")
            try:
                row_number = int(row_number_raw)
            except (TypeError, ValueError):
                continue

            blocker_row_keys.add((source_file, row_number))

            field_name = str(error.get("field", "")).strip()
            reason = str(error.get("reason", "")).strip()
            if field_name and reason:
                reasons_by_row[(source_file, row_number)].append(f"{field_name}: {reason}")

        for row in all_rows:
            row_key = (row["source_file"], row["row_number"])
            if row_key not in blocker_row_keys:
                continue

            row_reasons = reasons_by_row.get(row_key, [])
            unique_reasons = list(dict.fromkeys(row_reasons))

            for field_name, value in row.get("data", {}).items():
                normalized = comparable(value)
                if normalized is None:
                    continue
                index[(entity_name, str(field_name))][normalized].append(
                    {
                        "source_file": row["source_file"],
                        "row_number": row["row_number"],
                        "reasons": unique_reasons,
                    }
                )

    return {
        key: {value: list(rows) for value, rows in value_map.items()}
        for key, value_map in index.items()
    }


def _serialize_row(raw_row: dict[str, Any]) -> str:
    return json.dumps(raw_row, ensure_ascii=False, default=str, sort_keys=True)


def _build_error_row(
    entity_name: str,
    source_file: str,
    row_number: int,
    field_name: str,
    reason: str,
    original_value: Any,
    raw_row: dict[str, Any],
) -> dict[str, Any]:
    dependency_meta = _parse_dependency_metadata(reason)
    _error_code, severity, suggested_fix = _classify_error(
        field_name=field_name,
        reason=reason,
        dependency_type=dependency_meta["dependency_type"],
    )

    return {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "entity": entity_name,
        "source_file": source_file,
        "row_number": row_number,
        "field": field_name,
        "severity": severity,
        "dependency_type": dependency_meta["dependency_type"],
        "depends_on_entity": dependency_meta["depends_on_entity"],
        "depends_on_source_file": dependency_meta["depends_on_source_file"],
        "depends_on_row_number": dependency_meta["depends_on_row_number"],
        "suggested_fix": suggested_fix,
        "reason": reason,
        "original_value": "" if original_value is None else str(original_value),
        "raw_row": _serialize_row(raw_row),
    }


def _write_csv(file_path: Path, rows: list[dict[str, Any]], fieldnames: list[str]):
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with file_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def _append_run_log(log_path: Path, summaries: list[dict[str, Any]]):
    log_path.parent.mkdir(parents=True, exist_ok=True)
    file_exists = log_path.exists()
    with log_path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=RUN_LOG_FIELDS)
        if not file_exists:
            writer.writeheader()

        timestamp = datetime.now().isoformat(timespec="seconds")
        for summary in summaries:
            writer.writerow(
                {
                    "timestamp": timestamp,
                    "entity": summary["entity"],
                    "files_processed": "; ".join(summary["files_processed"]),
                    "rows_read": summary["rows_read"],
                    "rows_accepted": summary["rows_accepted"],
                    "rows_rejected": summary["rows_rejected"],
                    "duplicate_rows_flagged": summary.get("duplicate_rows_flagged", 0),
                    "output_file": summary["output_file"],
                    "error_file": summary["error_file"],
                }
            )


def _append_duplicate_log(log_path: Path, rows: list[dict[str, Any]]):
    log_path.parent.mkdir(parents=True, exist_ok=True)
    file_exists = log_path.exists()
    with log_path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DUPLICATE_LOG_FIELDS)
        if not file_exists:
            writer.writeheader()

        for row in rows:
            writer.writerow({field: row.get(field, "") for field in DUPLICATE_LOG_FIELDS})
